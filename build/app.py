import subprocess
import csv
import sys
import random

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", bs4])
    subprocess.check_call([sys.executable, "-m", "pip", "install", requests])
    subprocess.check_call([sys.executable, "-m", "pip", "install", datetime])

from bs4 import BeautifulSoup
from tkinter import *
from tkinter import ttk
from tkinter import font
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import datetime as dt


app = Tk()
app.geometry('450x300')
app.title("TTSF Cloud One")
app.iconbitmap(r'build\app\icons\ttsf.ico')
app['background']="MediumPurple4"

#creating a mainframe
main_frame = Frame(app)
main_frame.pack(fill=BOTH, expand=1)
main_frame['background']='MediumPurple4'

#creating a canvas
my_canvas = Canvas(main_frame)
my_canvas.pack(side=LEFT, fill=BOTH, expand=1)
my_canvas['background']='MediumPurple4'

#add scrollbar to the Canvas
my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
my_scrollbar.pack(side=RIGHT, fill=Y)

#configure the canvas
my_canvas.configure(yscrollcommand=my_scrollbar.set)
my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion = my_canvas.bbox("all")))

#create another frame inside canvas
second_frame = Frame(my_canvas)
second_frame['background']='MediumPurple4'

#Add new frame to the window in the canvas
my_canvas.create_window((0,0), window=second_frame, anchor="nw")



msg=Label(second_frame,text="  TTSF Cloud One",background="MediumPurple4",foreground="darkorange1", font=("Times",32, "bold"))
msg.grid(row=0,column=0,padx=25,pady=20,columnspan=6)



#linksE.grid(row=2,column=1,padx=50,pady=15)
#r = Label(second_frame, text=linksE.get())                          #getting all elements of the location (Home/India/Bangalore/HSR/)
#r.grid(row=4,column=1,padx=50,pady=15)


class restaurant:              #Creating a class called restaurant
    def __init__(self, link):  #initializing the link and header file
        self.link = link
        lnk='https://www.swiggy.com/restaurants/prowl-foods-by-tiger-shroff-indira-nagar-indiranagar-bangalore-405907'
        self.hdr = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0',
                    'Accept-Language': 'en-GB,enl;q=0.5', 'Referer': 'https://google.com',
                    'DNT': '1'}  # header used to replicate as a user request rather than python request while accesing websites
        try:
            html = requests.get(self.link, headers=self.hdr)
        except Exception:
            html = requests.get(lnk, headers=self.hdr)
            self.link = lnk
        self.soup = BeautifulSoup(html.text, 'html.parser')


    def rest_name(self):  # to get the name of the restaurant
        if self.link[12:15] == "swi":  # to check if link is of Swiggy
            name = self.soup.find('span', class_='kpkwa')
            self.n1=name.text
            ln = '(Swiggy)'
            return name.text+'  '+ln  # returns name of
        elif self.link[12:15] == "zom":  # to check if link is of Zomato
            name = self.soup.find('h1', class_='sc-7kepeu-0 sc-kafWEX kTxZkY')
            self.n1 = name.text
            ln = '(Zomato)'
            return name.text+'  '+ln

    def location(self):  # to get the location of the restaurant
        i=0
        if self.link[12:15] == "swi":   #to check if link is of wiggy
            locate = self.soup.find_all('span', class_='_3duMr')
            loc = []   #assigning empty list
            for l in locate:   #iterating through all span elements
                loc.append(l.text)  # getting all elements of the location (Home/India/Bangalore/HSR/)
            return loc[-1]  # the last element is always the location of the restaurant
        elif self.link[12:15] == "zom":  # if the given link is of zomato following lines are executed
            location = self.soup.find_all('span', class_='sc-ks3f96-1 gETRUR')
            loc = []    #assigning empty list
            for l in location:    #iterating through all span elements
                loc.append(l.text)  # getting all elements of the location (Home/India/Bangalore/HSR/)
            return (loc[4][:len(loc[4]) - 1]) #usually the 4th element is the location of the restaurant'''



    def ratings(self):  # function to get the ratings of the restaurant
        if self.link[12:15] == "swi":   # if the given link is of Swiggy following lines are executed
            Ratings = self.soup.find('div', class_='_1BpLF')
            rate = []   #assigning empty list
            for r in Ratings:    #iterating through all span elements
                rate.append(r.text)  # getting all the text from the ratings block
            return (f'{rate[-1][:3]}⋆')  # the last element is always the location of the restaurant
        elif self.link[12:15] == "zom":  # if the given link is of zomato following lines are executed
            Ratings = self.soup.find_all('div', class_='sc-1q7bklc-5 clCBXa')
            rate = []     #assigning empty list
            for r in Ratings:    #iterating through all span elements
                rate.append(r.text)  # getting all the text from the ratings block
            return (f'{rate[-1][:3]}⋆')   # the last element is the location sliced till 3 elements



    def sections(self):                         # to get the categories sections (left panel)
        if self.link[12:15] == "swi":                   # if the given link is of Swiggy following lines are executed
            section = self.soup.find_all('h2', class_='M_o7R')         #all the h2 tags are the categories
            self.sec = []                  #creating an empty list called sec
            for sect in section:       #iterating through all h2 tags
                self.sec.append(sect.text)  # making a list of all the h2 tags
            return self.sec
        elif self.link[12:15] == "zom":       # if the given link is of Zomato following lines are executed
            section = self.soup.find_all('p')  # gettiing all the p tags
            self.sec1 = []       #assigning empty list
            sec2 = []  #assigning empty list
            sec3 = []  #assigning empty list
            self.sec = []  #assigning empty list
            for sect in section:   #iterating through all p tags
                sec2.append(sect.text)  # appending all the p tag elements to the list k
            for i in sec2:
                if i != '':  # checking if any empty string elements are present in the list
                    sec3.append(i)  # removing unwanted data from the list and appending it
            for j in range(len(sec3)):
                if sec3[j][-1] == ')':  # Recommended (8) #checking if the last element is ')' then appending it to a list
                    self.sec1.append(sec3[j])
            for p in self.sec1:
                j = p.split('(')       # splitting the category so that it only has the element [Pastas] instead of [Pastas (8)]
                self.sec.append(j[0][:-1])  #slicing till last element to remove the empty spaces in the end
            return self.sec  # Returns the list of all sections


    def products(self):
        if self.link[12:15] == "swi":  # if the given link is of Swiggy following lines are executed
            products = self.soup.find_all('div', class_='styles_itemName__hLfgz')
            self.productnames = []   #assigning an empty list
            for product in products:   #iterating through all the div tags
                self.productnames.append(product.text)  # holds all the product names
            return self.productnames
        elif self.link[12:15] == "zom":  # if the given link is of Zomato following lines are executed
            products = self.soup.find_all('h4', class_='sc-1s0saks-15 iSmBPS')
            self.productnames = []     #assigning an empty list
            for product in products:    #iterating through all the h4 tags
                self.productnames.append(product.text)  # holds all the product names
            return self.productnames


    def menu(self):  # to return a dictionary containing all the sections and products
        if self.link[12:15] == "swi":  #checking if the link is of swiggy
            section = self.soup.find_all('div', class_='_2dS-v')
            category = []   #empty list to hold all categories
            element = []   #empty list to hold all products
            for sec in section:   #iterating through all div tags
                title = sec.find('h2').text  # to get the section names
                products = sec.find_all('h3', class_="styles_itemNameText__3ZmZZ")  # to get all the product names
                pro = []   #empty list to assign products
                for prod in products:  #iterating through the products
                    pro.append(prod.text)   #appending all the products
                category.append(title)   #appending titles
                element.append(pro)  #appending lists of products
            self.menu1 = dict(zip(category, element))  # Joining categories and products make them a dictionary
            return self.menu1
        elif self.link[12:15] == "zom":   #checking if the link is of Zomato
            self.sections()     #calling another function sections()
            self.products()     #calling another function products()
            product_no = []    #assigning empty lists
            prod = []    #assigning empty lists
            for i in self.sec1:  # contains the category list from the previous function (sections)
                k = i.split('(')   #splitting the element
                if (k[-1][:-1]).isdigit():
                    product_no.append(int(k[-1][:-1]))   #appending all the category numbers
                else:
                    try:
                        sp=k[-1][:-1].slit('(')
                        product_no.append(int(sp[-1][:-1]))
                    except Exception:
                        pass

            prod_name = list(self.productnames)  # contains the product names from the previous function (products)
            for j in product_no:     #iterating through all product numbers
                product_list = []    #empty list
                for k in range(j):
                    product_list.append(prod_name[0])   #appending all the product names
                    prod_name.pop(0)     #first element is the section name (to delete that)
                prod.append(product_list)  #appending list of product list
            self.menu1 = dict(zip(self.sec, prod))  # Joining categories and products make them a dictionary
            return self.menu1

    def unavail(self):
        if self.n1.lower()=='the thick shake factory':
            self.n1 ='the thickshake factory'
            print(self.n1.lower())
        if self.n1.lower()=='prowl foods by tiger shroff':
            prowl_list=[]
            prowl_list1=[]
            prowl_cat=[]
            with open('Menu\\prowl.csv','r') as prowl:
                prowl_r = csv.reader(prowl)
                for line in prowl_r:
                    prowl_list.append(line)
            for lst in prowl_list:
                prowl_cat.append(lst[0])
                lst.pop(0)

                prowl_list1.append(lst)

            prowl_dict = dict(zip(prowl_cat,prowl_list1))
            wp_dict = self.menu1
            unavail_key = []
            unavail_prod1 = []
            unavail_prod = []
            for key1 in prowl_dict:
                if key1 not in list(wp_dict.keys()):
                    wp_dict[key1]=''
                for key in wp_dict:
                    if key1==key:
                        unavail_key.append(key)
                        for items in prowl_dict[key]:
                            if items not in wp_dict[key1]:
                                unavail_prod.append(items)
                        unavail_prod1.append(unavail_prod)
                        unavail_prod=[]
                    else:
                        pass

            unavail_products = dict(zip(unavail_key,unavail_prod1))

            #for i in unavail_products:
            #    print(i)
            #    print(unavail_products[i])
            return unavail_products

        if self.n1.lower() == 'the thickshake factory':
            ttsf_list = []
            ttsf_list1 = []
            ttsf_cat = []
            with open('Menu\\ttsf.csv', 'r') as ttsf:
                ttsf_r = csv.reader(ttsf)
                for line in ttsf_r:
                    ttsf_list.append(line)
            for lst in ttsf_list:
                ttsf_cat.append(lst[0])
                lst.pop(0)

                ttsf_list1.append(lst)

            ttsf_dict = dict(zip(ttsf_cat, ttsf_list1))
            wp_dict = self.menu1
            unavail_key = []
            unavail_prod1 = []
            unavail_prod = []
            for key1 in ttsf_dict:
                if key1 not in list(wp_dict.keys()):
                    wp_dict[key1] = ''
                for key in wp_dict:
                    if key1 == key:
                        unavail_key.append(key)
                        for items in ttsf_dict[key]:
                            if items not in wp_dict[key1]:
                                unavail_prod.append(items)
                        unavail_prod1.append(unavail_prod)
                        unavail_prod = []
                    else:
                        pass

            unavail_products = dict(zip(unavail_key, unavail_prod1))

            #for i in unavail_products:
            #    print(i)
            #    print(unavail_products[i])
            return unavail_products

        if self.n1.lower() == 'wowffles':
            wow_list = []
            wow_list1 = []
            wow_cat = []
            with open('Menu\\wowffles.csv', 'r') as wow:
                wow_r = csv.reader(wow)
                for line in wow_r:
                    wow_list.append(line)
            for lst in wow_list:
                wow_cat.append(lst[0])
                lst.pop(0)

                wow_list1.append(lst)

            wow_dict = dict(zip(wow_cat, wow_list1))
            wp_dict = self.menu1
            unavail_key = []
            unavail_prod1 = []
            unavail_prod = []
            for key1 in wow_dict:
                if key1 not in list(wp_dict.keys()):
                    wp_dict[key1] = ''
                for key in wp_dict:
                    if key1 == key:
                        unavail_key.append(key)
                        for items in wow_dict[key]:
                            if items not in wp_dict[key1]:
                                unavail_prod.append(items)
                        unavail_prod1.append(unavail_prod)
                        unavail_prod = []
                    else:
                        pass

            self.unavail_products = dict(zip(unavail_key, unavail_prod1))

            #for i in unavail_products:
            #    print(i)
            #    print(unavail_products[i])
            return self.unavail_products

        else:
            return self.menu1



def excel():
    lnks = []
    with open('links.txt', 'r') as f:
        for line in f:
            if line[-1] == '\n':
                lnks.append(line[:-1])
            else:
                lnks.append(line)

    wb = Workbook()
    ws = wb.active
    ws.title = "Store Data"
    ws.merge_cells('E1:J2')
    ws['E1'].value = 'TTSF Cloud One'
    ws['E1'].font = Font(name='Arial', bold=True, size=22, color='00660066', underline="single")
    ws['E1'].alignment = Alignment(horizontal="center", vertical="center")
    j = '3'

    for i in range(len(lnks)):
        obj1 = restaurant(lnks[i])

        if lnks[i][12:15] == "swi":
            link = '(Swiggy)'
        else:
            link = '(Zomato)'
        name = obj1.rest_name()  # to get the Restaurant name
        locate = obj1.location()  # to find the location
        rate = obj1.ratings()  # to find the ratings
        menu_dict = obj1.menu()  # to get a dictionary with categories and products
        if list(menu_dict.keys())[0] == 'Recommended':
            del menu_dict['Recommended']

        ws['A' + j] = name.split('(')[0][:-1]
        ws['A' + j].font = Font(name='Arial', bold=True, size=16, color='00666699', underline='single')
        #ws['A' + j].alignment = Alignment(horizontal="center", vertical="center")
        ws['C' + j] = link
        ws['C' + j].font = Font(name='Arial', bold=True, size=10, color='00339966', italic=True)
        ws['C' + j].alignment = Alignment(horizontal="center", vertical="center")
        j = str(int(j) + 1)

        ws['A' + j] = 'Location :'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080')
        ws['B' + j] = locate
        ws['B' + j].font = Font(size=12, color='00666699')
        j = str(int(j) + 1)

        ws['A' + j] = 'Ratings :'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080')
        ws['B' + j] = rate
        ws['B' + j].font = Font(size=12, color='00666699')
        j = str(int(j) + 2)

        ws['A' + j] = 'Categories with Products'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080', italic=True, underline='single')
        j = str(int(j) + 1)

        for key in menu_dict:
            category_no = len(menu_dict[key])
            k = key + ' (' + str(category_no) + ') ' + ' : '
            products = list(menu_dict[key])
            ws.append([k] + products)
            ws['A' + j].font = Font(bold=True, size=12, color='00993300')
            j = str(int(j) + 1)
        j = str(int(j) + 3)

        date = str(dt.datetime.now()).split()[0]
        time = dt.datetime.now().strftime("(%H_%M_%S)")
        file_name = 'Excel_files\\'+(date+time)+'.xlsx'

    ws.column_dimensions['A'].width = 40
    ws.sheet_view.showGridLines = False
    wb.save(file_name)


def excel_unavailble():
    lnks = []
    with open('links.txt', 'r') as f:
        for line in f:
            if line[-1] == '\n':
                lnks.append(line[:-1])
            else:
                lnks.append(line)

    wb = Workbook()
    ws = wb.active
    ws.title = "Store Data"
    ws.merge_cells('E1:J2')
    ws['E1'].value = 'TTSF Cloud One'
    ws['E1'].font = Font(name='Arial', bold=True, size=22, color='00660066', underline="single")
    ws['E1'].alignment = Alignment(horizontal="center", vertical="center")
    j = '3'

    for i in range(len(lnks)):
        obj2 = restaurant(lnks[i])

        if lnks[i][12:15] == "swi":
            link = '(Swiggy)'
        else:
            link = '(Zomato)'
        name = obj2.rest_name()  # to get the Restaurant name
        locate = obj2.location()  # to find the location
        rate = obj2.ratings()  # to find the ratings
        menu_dict = obj2.menu()  # to get a dictionary with categories and products
        unavil_dict = obj2.unavail()
        if list(unavil_dict.keys())[0] == 'Recommended':
            del unavil_dict['Recommended']

        ws['A' + j] = name.split('(')[0][:-1]
        ws['A' + j].font = Font(name='Arial', bold=True, size=16, color='00666699', underline='single')
        #ws['A' + j].alignment = Alignment(horizontal="center", vertical="center")
        ws['C' + j] = link
        ws['C' + j].font = Font(name='Arial', bold=True, size=10, color='00339966', italic=True)
        ws['C' + j].alignment = Alignment(horizontal="center", vertical="center")
        j = str(int(j) + 1)

        ws['A' + j] = 'Location :'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080')
        ws['B' + j] = locate
        ws['B' + j].font = Font(size=12, color='00666699')
        j = str(int(j) + 1)

        ws['A' + j] = 'Ratings :'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080')
        ws['B' + j] = rate
        ws['B' + j].font = Font(size=12, color='00666699')
        j = str(int(j) + 2)

        ws['A' + j] = 'Categories with Products'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080', italic=True, underline='single')
        j = str(int(j) + 1)

        for key in unavil_dict:
            category_no = len(unavil_dict[key])
            k = key + ' (' + str(category_no) + ') ' + ' : '
            products = list(unavil_dict[key])
            ws.append([k] + products)
            ws['A' + j].font = Font(bold=True, size=12, color='00993300')
            j = str(int(j) + 1)
        j = str(int(j) + 3)

        date = str(dt.datetime.now()).split()[0]
        time = dt.datetime.now().strftime("(%H_%M_%S)")
        file_name = 'Excel_files_unavailable\\'+(date+time)+'.xlsx'

    ws.column_dimensions['A'].width = 40
    ws.sheet_view.showGridLines = False
    wb.save(file_name)



def button_on(e):
    B["bg"] = "thistle1"

def button_off(e):
    B["bg"] = "thistle2"

def button_on1(e):
    B1["bg"] = "thistle1"

def button_off1(e):
    B1["bg"] = "thistle2"

def button_on2(e):
    B2["bg"] = "thistle1"

def button_off2(e):
    B2["bg"] = "thistle2"



def unavailable():
    app1 = Toplevel()
    app1.geometry('900x700')
    app1.title("TTSF")
    app1.iconbitmap(r'build\app\icons\ttsf.ico')
    app1['background'] = "MediumPurple4"

    # creating a mainframe
    main_frame = Frame(app1)
    main_frame.pack(fill=BOTH, expand=1)
    main_frame['background'] = 'MediumPurple4'

    # creating a canvas
    my_canvas = Canvas(main_frame)
    my_canvas.pack(side=LEFT, fill=BOTH, expand=1)
    my_canvas['background'] = 'MediumPurple4'

    # add scrollbar to the Canvas
    my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
    my_scrollbar.pack(side=RIGHT, fill=Y)

    # configure the canvas
    my_canvas.configure(yscrollcommand=my_scrollbar.set)
    my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))

    # create another frame inside canvas
    third_frame = Frame(my_canvas)
    third_frame['background'] = 'MediumPurple4'

    # Add new frame to the window in the canvas
    my_canvas.create_window((0, 0), window=third_frame, anchor="nw")

    resta = Label(third_frame, text="Unavailable Products", background="MediumPurple4", foreground="wheat1",font=("Times", 32, "bold"))
    resta.grid(row=0, column=0, padx=25, pady=20, columnspan=7)

    B3 = Button(third_frame, text="Export to Excel", command=excel_unavailble, background='thistle2', width=13)
    B3.grid(row=1, column=2, padx=15, pady=25, columnspan=1)

    links = []
    with open('links.txt', 'r') as f:
        for line in f:
            if line[-1] == '\n':
                links.append(line[:-1])
            else:
                links.append(line)
    #links=links[14:15]

    row_val = 2
    for i in range(len(links)):

        obj1 = restaurant(links[i])  # initializing an object to the class restaurant()

        name = obj1.rest_name()  # to get the Restaurant name
        # p = Label(second_frame, text='Restaurant name : ', background="MediumPurple4", foreground="Dark Slate gray",
        #          font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        # p.grid(row=row_val, column=0, padx=10, pady=3)
        p = Label(third_frame, text=name, bd=1, background="MediumPurple4", foreground="light salmon",
                  font=("Vijaya", 22, "bold"))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=0, padx=5, pady=10, columnspan=2)
        row_val += 1

        locate = obj1.location()  # to find the location
        p = Label(third_frame, text='Location : ', background="MediumPurple4", foreground="Plum1",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=0, padx=10, pady=3)
        p = Label(third_frame, text=locate, background="MediumPurple4", foreground="light cyan",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=1, padx=10, pady=3)
        row_val += 1

        ratings = obj1.ratings()  # to get the ratings
        p = Label(third_frame, text='Ratings : ', background="MediumPurple4", foreground="Plum2",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=0, padx=10, pady=3)
        p = Label(third_frame, text=ratings, background="MediumPurple4", foreground="light cyan",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=1, padx=10, pady=3)
        row_val += 1

        p = Label(third_frame, text='Categories with Products', background="MediumPurple4", foreground="lavender",
                  font=(
                      "Verdana", 14, "italic",
                      "underline"))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=0, padx=5, pady=2)
        row_val += 1

        menu1 = obj1.menu()  # to get the menu
        # if list(menu.keys())[0] == 'Recommended':
        #    del menu['Recommended']
        menu = obj1.unavail()

        for sec in menu:
            p = Label(third_frame, text=sec + ' ' + '(' + str(len(menu[sec])) + ')', background="MediumPurple4",
                      foreground='Plum2',
                      font=("Verdana", 12))  # getting all elements of the location (Home/India/Bangalore/HSR/)
            p.grid(row=row_val, column=0, padx=2, pady=3)
            items = ','.join(menu[sec])
            l = Text(third_frame, width=50, height=1, background='thistle1', foreground='Black',
                     font=("Verdana", 10))
            l.insert(1.0, items)
            l.grid(row=row_val, column=1, padx=2, pady=3)
            row_val += 1



        p = Label(third_frame, text=' ', background="MediumPurple4",
                  foreground="lightsteelblue2",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=1, padx=10, pady=25)
        row_val += 1

        row_val += 2






def run():
    app.geometry('1050x680')
    resta = Label(second_frame, text=" Restaurants Data ", background="MediumPurple4", foreground="wheat1",font=("Verdana", 18, "bold", "underline"))
    resta.grid(row=2, column=0, padx=25, pady=10, columnspan=7)
    links = []
    with open('links.txt', 'r') as f:
        for line in f:
            if line[-1] == '\n':
                links.append(line[:-1])
            else:
                links.append(line)
    #links=links[:1]
    row_val = 4
    for i in range(len(links)):

        obj = restaurant(links[i])  # initializing an object to the class restaurant()

        name = obj.rest_name()  # to get the Restaurant name
        # p = Label(second_frame, text='Restaurant name : ', background="MediumPurple4", foreground="Dark Slate gray",
        #          font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        # p.grid(row=row_val, column=0, padx=10, pady=3)
        p = Label(second_frame, text=name, bd=1, background="MediumPurple4", foreground="light salmon",
                  font=("Vijaya", 22, "bold"))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=0, padx=5, pady=10, columnspan=2)
        row_val += 1

        locate = obj.location()  # to find the location
        p = Label(second_frame, text='Location : ', background="MediumPurple4", foreground="Plum1",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=0, padx=10, pady=3)
        p = Label(second_frame, text=locate, background="MediumPurple4", foreground="light cyan",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=1, padx=10, pady=3)
        row_val += 1

        ratings = obj.ratings()  # to get the ratings
        p = Label(second_frame, text='Ratings : ', background="MediumPurple4", foreground="Plum2",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=0, padx=10, pady=3)
        p = Label(second_frame, text=ratings, background="MediumPurple4", foreground="light cyan",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=1, padx=10, pady=3)
        row_val += 1

        p = Label(second_frame, text='Categories with Products', background="MediumPurple4", foreground="lavender", font=(
        "Verdana", 14, "italic", "underline"))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=0, padx=5, pady=2)
        row_val += 1

        menu = obj.menu()  # to get the menu
        # if list(menu.keys())[0] == 'Recommended':
        #    del menu['Recommended']
        for sec in menu:
            p = Label(second_frame, text=sec + ' ' + '(' + str(len(menu[sec])) + ')', background="MediumPurple4",
                      foreground='Plum2',
                      font=("Verdana", 12))  # getting all elements of the location (Home/India/Bangalore/HSR/)
            p.grid(row=row_val, column=0, padx=2, pady=3)
            items = ','.join(menu[sec])
            l = Text(second_frame, width=50, height=1, background='thistle1', foreground='Black',
                     font=("Verdana", 10))
            l.insert(1.0, items)
            l.grid(row=row_val, column=1, padx=2, pady=3)
            row_val += 1

        p = Label(second_frame, text=' ', background="MediumPurple4",
                  foreground="lightsteelblue2",
                  font=("Verdana", 14))  # getting all elements of the location (Home/India/Bangalore/HSR/)
        p.grid(row=row_val, column=1, padx=10, pady=25)
        row_val += 1

        row_val += 2



hdr = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0',
       'Accept-Language' : 'en-GB,enl;q=0.5',
       'Referer' : 'https://google.com',
       'DNT' : '1'}                                              #header used to replicate as a user request rather than python request while accesing websites




B=Button(second_frame, text="Export to Excel",command=excel, background='thistle2',width=13)
B.grid(row=1,column=4,padx=15,pady=25, columnspan=2)

B1=Button(second_frame, text="Run",command=run, background='thistle2', width=13)
B1.grid(row=1,column=0,padx=15,pady=25, columnspan=1)

B2=Button(second_frame, text="Unavailable",command=unavailable, background='thistle2', width=13)
B2.grid(row=1,column=2,padx=15,pady=25, columnspan=1)

B.bind("<Enter>", button_on)
B.bind("<Leave>", button_off)

B1.bind("<Enter>", button_on1)
B1.bind("<Leave>", button_off1)

B2.bind("<Enter>", button_on2)
B2.bind("<Leave>", button_off2)

#if __name__ == "__main__":
mainloop()